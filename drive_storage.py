"""
Google Drive XLSX Storage Utility

Stores data as .xlsx files in Google Drive using OAuth credentials.
Data is organized into Latest (overwritten) and Cumulative (upserted) folders.

Folder structure:
  GOOGLE_DRIVE_FOLDER_ID/
    2026/
      Latest/
        news_raw.xlsx
        download_rank_7d.xlsx
        ...
      Cumulative/
        news_raw.xlsx
        download_rank_7d.xlsx
        ...

Required environment variables:
  GOOGLE_OAUTH_CLIENT_ID      — OAuth client ID
  GOOGLE_OAUTH_CLIENT_SECRET  — OAuth client secret
  GOOGLE_OAUTH_REFRESH_TOKEN  — OAuth refresh token
  GOOGLE_DRIVE_FOLDER_ID      — Root folder ID
"""

import os
import io
import logging
from datetime import datetime, timezone

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaInMemoryUpload
from openpyxl import Workbook, load_workbook

log = logging.getLogger(__name__)

SCOPES = ["https://www.googleapis.com/auth/drive"]
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FOLDER_MIME = "application/vnd.google-apps.folder"

_service = None


def get_drive_service():
    """Authenticate via OAuth refresh token and return a Drive API service (cached)."""
    global _service
    if _service:
        return _service

    client_id = os.environ.get("GOOGLE_OAUTH_CLIENT_ID", "")
    client_secret = os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET", "")
    refresh_token = os.environ.get("GOOGLE_OAUTH_REFRESH_TOKEN", "")

    if not all([client_id, client_secret, refresh_token]):
        raise RuntimeError(
            "Missing Google OAuth env vars. Need: "
            "GOOGLE_OAUTH_CLIENT_ID, GOOGLE_OAUTH_CLIENT_SECRET, GOOGLE_OAUTH_REFRESH_TOKEN"
        )

    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=client_id,
        client_secret=client_secret,
        scopes=SCOPES,
    )
    creds.refresh(Request())

    _service = build("drive", "v3", credentials=creds, cache_discovery=False)
    return _service


def get_folder_id():
    """Get the root Google Drive folder ID from environment."""
    folder_id = os.environ.get("GOOGLE_DRIVE_FOLDER_ID", "")
    if not folder_id:
        raise RuntimeError("GOOGLE_DRIVE_FOLDER_ID is not set")
    return folder_id


# ─── Folder management ──────────────────────────────────────────────────────

def ensure_subfolder(parent_id, folder_name):
    """Find or create a subfolder under parent_id. Returns the folder ID."""
    service = get_drive_service()

    query = (f"name = '{folder_name}' and '{parent_id}' in parents "
             f"and mimeType = '{FOLDER_MIME}' and trashed = false")
    results = service.files().list(q=query, fields="files(id)", pageSize=1).execute()
    files = results.get("files", [])

    if files:
        return files[0]["id"]

    metadata = {
        "name": folder_name,
        "parents": [parent_id],
        "mimeType": FOLDER_MIME,
    }
    folder = service.files().create(body=metadata, fields="id").execute()
    print(f"  Created folder: {folder_name}")
    return folder["id"]


def _get_year_folder(year=None):
    """Get (or create) the year subfolder under root. Returns folder ID."""
    if year is None:
        year = datetime.now(timezone.utc).year
    root = get_folder_id()
    return ensure_subfolder(root, str(year))


def get_latest_folder(year=None):
    """Get (or create) the Latest folder: root/{year}/Latest."""
    year_folder = _get_year_folder(year)
    return ensure_subfolder(year_folder, "Latest")


def get_cumulative_folder(year=None):
    """Get (or create) the Cumulative folder: root/{year}/Cumulative."""
    year_folder = _get_year_folder(year)
    return ensure_subfolder(year_folder, "Cumulative")


def _latest_filename(base_name, year=None):
    """Generate Latest filename: news_raw -> news_raw_2026_latest.xlsx"""
    if year is None:
        year = datetime.now(timezone.utc).year
    base = base_name.removesuffix(".xlsx").removesuffix(".csv")
    return f"{base}_{year}_latest.xlsx"


def _cumulative_filename(base_name, year=None):
    """Generate Cumulative filename: news_raw -> news_raw_2026.xlsx"""
    if year is None:
        year = datetime.now(timezone.utc).year
    base = base_name.removesuffix(".xlsx").removesuffix(".csv")
    return f"{base}_{year}.xlsx"


# ─── File operations ─────────────────────────────────────────────────────────

def find_file_in_folder(filename, folder_id):
    """Find a file by name in a specific folder. Returns file ID or None."""
    service = get_drive_service()
    query = f"name = '{filename}' and '{folder_id}' in parents and trashed = false"
    results = service.files().list(q=query, fields="files(id, name)", pageSize=1).execute()
    files = results.get("files", [])
    return files[0]["id"] if files else None


def _rows_to_xlsx_bytes(rows, headers):
    """Convert a list of dicts to XLSX bytes."""
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_xlsx_by_id(file_id):
    """Read an XLSX file by file ID, return list of dicts."""
    service = get_drive_service()
    content = service.files().get_media(fileId=file_id).execute()
    if not content:
        return []
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    return [{h: (cell if cell is not None else "") for h, cell in zip(headers, row)} for row in rows_iter]


def _write_xlsx_to_folder(filename, rows, headers, folder_id):
    """Write rows as an XLSX file in the specified folder (create or update)."""
    service = get_drive_service()
    content = _rows_to_xlsx_bytes(rows, headers)
    media = MediaInMemoryUpload(content, mimetype=XLSX_MIME)

    existing_id = find_file_in_folder(filename, folder_id)

    if existing_id:
        service.files().update(fileId=existing_id, media_body=media).execute()
        print(f"  Updated {filename} ({len(rows)} rows)")
    else:
        metadata = {
            "name": filename,
            "parents": [folder_id],
            "mimeType": XLSX_MIME,
        }
        service.files().create(body=metadata, media_body=media).execute()
        print(f"  Created {filename} ({len(rows)} rows)")


# ─── High-level storage functions ────────────────────────────────────────────

def save_latest_and_cumulative(base_filename, rows, headers, dedup_keys):
    """Two-step save: overwrite Latest, then upsert into Cumulative.

    Step 1 — Latest: Overwrite the file with only the new rows.
    Step 2 — Cumulative: Read existing data, merge with new rows
             deduplicating by dedup_keys (new rows win on conflict).

    Args:
        base_filename: File name (e.g. "download_rank_7d.xlsx")
        rows: List of dicts to store
        headers: Column names
        dedup_keys: List of field names that together form a unique key
                    (should NOT include volatile fields like datetime_of_news)
    """
    if not rows:
        print(f"  No rows to save for {base_filename}")
        return 0

    latest_name = _latest_filename(base_filename)
    cumulative_name = _cumulative_filename(base_filename)

    # Step 1: Overwrite Latest
    latest_folder = get_latest_folder()
    _write_xlsx_to_folder(latest_name, rows, headers, latest_folder)
    print(f"  [Latest] Wrote {len(rows)} rows to {latest_name}")

    # Step 2: Upsert into Cumulative
    cumulative_folder = get_cumulative_folder()
    file_id = find_file_in_folder(cumulative_name, cumulative_folder)

    if file_id:
        existing = _read_xlsx_by_id(file_id)
        # Build a set of composite keys from the new rows
        def _make_key(row):
            return tuple(str(row.get(k, "")) for k in dedup_keys)

        new_keys = {_make_key(r) for r in rows}
        # Keep existing rows whose key is NOT in the new batch
        filtered = [r for r in existing if _make_key(r) not in new_keys]
        replaced = len(existing) - len(filtered)
        all_rows = filtered + rows
        if replaced:
            print(f"  [Cumulative] Replacing {replaced} existing rows")
    else:
        all_rows = rows

    _write_xlsx_to_folder(cumulative_name, all_rows, headers, cumulative_folder)
    print(f"  [Cumulative] Saved {len(rows)} new rows to {cumulative_name} (total: {len(all_rows)})")
    return len(rows)


# ─── Read helpers ─────────────────────────────────────────────────────────────

def read_latest(base_filename, year=None):
    """Read an XLSX file from the Latest folder. Returns list of dicts."""
    folder_id = get_latest_folder(year)
    filename = _latest_filename(base_filename, year)

    file_id = find_file_in_folder(filename, folder_id)
    if not file_id:
        print(f"  {filename} not found in Latest")
        return []

    rows = _read_xlsx_by_id(file_id)
    print(f"  Read {len(rows)} rows from Latest/{filename}")
    return rows


def read_cumulative(base_filename, year=None):
    """Read an XLSX file from the Cumulative folder. Returns list of dicts."""
    folder_id = get_cumulative_folder(year)
    filename = _cumulative_filename(base_filename, year)

    file_id = find_file_in_folder(filename, folder_id)
    if not file_id:
        print(f"  {filename} not found in Cumulative")
        return []

    rows = _read_xlsx_by_id(file_id)
    print(f"  Read {len(rows)} rows from Cumulative/{filename}")
    return rows


# ─── Legacy helpers (root folder) ────────────────────────────────────────────

def find_file(filename):
    """Find a file by name in the root folder. Returns file ID or None."""
    return find_file_in_folder(filename, get_folder_id())


def read_xlsx(filename):
    """Read XLSX from root folder."""
    file_id = find_file(filename)
    if not file_id:
        print(f"  {filename} not found in Drive")
        return []
    rows = _read_xlsx_by_id(file_id)
    print(f"  Read {len(rows)} rows from {filename}")
    return rows


def upload_xlsx(filename, rows, headers):
    """Upload XLSX to the root folder."""
    _write_xlsx_to_folder(filename, rows, headers, get_folder_id())
    return len(rows)
